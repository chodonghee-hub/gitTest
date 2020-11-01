import openpyxl as op
import matplotlib.pyplot as plt

file_path = 'C:/Users/cho donghee/Desktop/board_data/'
user_data = 'user_list.xlsx'
article_data = 'article_list.xlsx'

def __loadUserList__():
    user_list = []
    wb = op.load_workbook(file_path + user_data)
    ws = wb['회원정보']
    for data in ws.rows :
        new_user = {'아이디':data[0].value, '비밀번호':data[1].value.strip(), '이름':data[2].value}
        user_list.append(new_user)
    return user_list

def __loadArticleList__():
    article_list = []
    wb = op.load_workbook(file_path + article_data)
    ws = wb['게시물정보']
    for data in ws.rows:
        if data[0].value == 'last_article_id': last_article_id = data[1].value
        else :
            new_article = {'번호':data[0].value, '제목':data[1].value, '내용':data[2].value, '작성자':data[3].value, '추천':data[4].value}
            article_list.append(new_article)
    return article_list, last_article_id

def __signUp__(new_id, new_pw, new_name, user_list):
    for user in user_list :
        if user['아이디'] == new_id : return False
    wb = op.load_workbook(file_path + user_data)
    sheet = wb.active
    sheet.append([new_id, new_pw, new_name])
    wb.save(file_path + user_data)
    return True

def __login__(enter_id, enter_pw, user_list):
    for user in user_list :
        if user['아이디'] == enter_id :
            if user['비밀번호'] == enter_pw :
                print('{}님 반갑습니다!!!'.format(user['이름']))
                return user
            else :
                print('비밀번호가 틀렸습니다.')
                return False
        elif user == user_list[-1] :
            print('없는 아이디 입니다.')
            return False

def read_article(article):
    print('\n● 게시물번호 : {} \n-- 제목 : {} \n-- 내용 : {} \n-- 작성자 : {} \n-- 추천 : {}'
          .format(article['번호'], article['제목'], article['내용'], article['작성자'], len(article['추천'].split('+')) if article['추천'] else 0))

# read version 2
def read_article_title(article):
    print('●번호 : {}  제목 : {}'.format(article['번호'], article['제목']))

def write_article(last_article_id, new_title, new_info, new_writer):
    wb = op.load_workbook(file_path + article_data)
    sheet = wb.active
    sheet.append([last_article_id+1, new_title, new_info, new_writer])
    sheet['B1'] = last_article_id+1
    wb.save(file_path + article_data)
    return __loadArticleList__()

def is_my_article(user, article):
    if user['이름'] == article['작성자'] : return True
    return False

def is_exists_article(article_list, article_no):
    for article in article_list :
        if article_no == article['번호'] : return article
    return False

def update_article(article_list, article, upd_key, upd_val):
    wb = op.load_workbook(file_path + article_data)
    sheet = wb.active
    row = article_list.index(article)
    article[upd_key] = upd_val

    sheet.cell(row=row+2, column=2).value = article['제목']
    sheet.cell(row=row+2, column=3).value = article['내용']

    wb.save(file_path + article_data)
    return __loadArticleList__()

def delete_article(article_list, article):
    wb = op.load_workbook(file_path + article_data)
    sheet = wb.active
    row = article_list.index(article)

    sheet.delete_rows(row+2)

    wb.save(file_path + article_data)
    return __loadArticleList__()

def chk_like_article(article_list, article, user):
    wb = op.load_workbook(file_path + article_data)
    sheet = wb.active
    row = article_list.index(article)

    like = article['추천'].split('+') if article['추천'] else []
    if user['이름'] not in like : like.append(user['이름'])
    else : like.remove(user['이름'])
    article['추천'] = '+'.join(like)

    sheet.cell(row=row+2, column=5).value = article['추천']
    wb.save(file_path + article_data)

def show_like_rate(article_list):
    likes = [len(article['추천'].split('+')) if article['추천'] else 0 for article in article_list]
    no_list = ['no : {}'.format(article['번호']) for article in article_list]

    plt.title('Like Rates')
    plt.xlabel('Article Number')
    plt.ylabel('Point')
    plt.bar(no_list, likes, width = 0.4, color = 'green')
    plt.show()

def start_pyBoard_login():
    print('===== Welcome to PyBoard =====')
    user_list = __loadUserList__()
    while True :
        print('==============================')
        print('[1] 로그인\n[2] 회원가입\n[3] 종료')
        cmd = input('▶ 명령어를 입력해주세요 : ')

        if cmd == '1' :
            print('\n▷ 로그인 모드')
            enter_id = input('-- 아이디를 입력해주세요 : ')
            enter_pw = input('-- 비밀번호를 입력해주세요 : ')
            rs = __login__(enter_id, enter_pw, user_list)
            if rs is False : continue
            else : start_pyBoard_main(rs)

        elif cmd == '2':
            print('\n▷ 회원가입 모드')
            enter_id = input('-- 희망하는 아이디를 입력해주세요 : ')
            enter_pw = input('-- 희망하는 비밀번호를 입력해주세요 : ')
            enter_name = input('-- 사용자의 이름을 입력해주세요 : ')
            rs = __signUp__(enter_id, enter_pw, enter_name, user_list)
            if rs :
                print('\n★ 회원가입 성공')
                user_list = __loadUserList__()
            else :
                print('\n☆ 회원가입 실패')

        elif cmd == '3':
            print('\n▷ 프로그램을 종료합니다')
            break

def start_pyBoard_main(user):
    article_list, last_article_id = __loadArticleList__()
    print('==============================')
    print('※ 접속중인 사용자 : {}'.format(user['이름']))
    while True :
        print('==============================')
        print('[1] 게시물 조회 \n[2] 게시물 작성 \n[3] 게시물 수정 \n[4] 게시물 삭제 \n[5] 게시물 추천 \n[6] 추천 통계 확인 \n[7] 로그아웃')
        cmd = input('▶ 명령어를 입력해주세요 : ')

        if cmd == '1':
            print('\n▷ 게시물조회 모드')
            select = input('[1] 제목조회 \n[2] 상세조회 \n>>> ')
            print('========= 게시물 목록 ==========')
            if select == '1':
                for article in article_list : read_article_title(article)
            elif select == '2' :
                for article in article_list : read_article(article)

        elif cmd == '2':
            print('\n▷ 게시물작성 모드')
            new_title = input('-- 제목을 입력해주세요 : ')
            new_info = input('-- 내용을 입력해주세요 : ')
            new_writer = user['이름']
            article_list, last_article_id = write_article(last_article_id, new_title, new_info, new_writer)

        elif cmd == '3':
            print('\n▷ 게시물수정 모드')
            article_no = int(input('-- 수정할 게시물 번호를 입력해주세요 : '))

            article = is_exists_article(article_list, article_no)
            if article != False:
                if is_my_article(user, article):
                    upd_key = input('-- 수정할 게시물 항목을 입력해주세요 : ')
                    upd_val = input('-- 수정할 게시물 내용을 입력해주세요 : ')
                    if upd_key != '번호' and upd_key != '작성자':
                        article_list, last_article_id = update_article(article_list, article, upd_key, upd_val)
                    else:
                        print("* '번호', '작성자'는 수정할 수 없습니다.")
                else:
                    print('* 게시물 수정 권한이 없습니다.')
            else : print('* 존재하지 않는 게시물 번호 입니다.')

        elif cmd == '4':
            print('\n▷ 게시물삭제 모드')
            article_no = int(input('-- 삭제할 게시물 번호를 입력해주세요 : '))

            article = is_exists_article(article_list, article_no)
            if article != False :
                if is_my_article(user, article):
                    article_list, last_article_id = delete_article(article_list, article)
                else : print('* 게시물 삭제 권한이 없습니다.')
            else : print('* 존재하지 않는 게시물 번호 입니다.')

        elif cmd == '5':
            print('▷ 게시물추천 모드')
            article_no = int(input('-- 추천할 게시물 번호를 입력해주세요 : '))
            article = is_exists_article(article_list, article_no)
            if article :
                chk_like_article(article_list, article, user)

        elif cmd == '6':
            print('▷ 추천통계 확인 모드')
            show_like_rate(article_list)

        elif cmd == '7':
            print('▷ 로그아웃 !!')
            break


start_pyBoard_login()