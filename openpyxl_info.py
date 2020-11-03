import openpyxl as op

file_path = 'C:/Users/SBS-/Desktop/dh/pyBoard/'

''' 
*** wb ( Workbook ) 필수 선언
*** ws ( WorkSheet ) 는 데이터 수정 (쓰기, 삭제) 일때는 wb.active // 데이터 읽기는 wb['Sheet_name'] 으로 초기화 
*** 데이터 수정 -> 쓰기 기능으로 셀 덮어쓰기 
*** 쓰기, 수정, 삭제 -- 데이터 수정 시 무조건 save(file_path + file_name)
'''


def create_new_excel():
    wb = op.Workbook()
    sheet = wb.active
    sheet.title = 'test_sheet'
    sheet['A1'] = 'test data'
    wb.save(file_path + 'test_sample.xlsx')

def read_excel(file_name):
    wb = op.load_workbook(file_path + file_name)
    ws = wb['test_sheet']

    for row in ws.rows:
        for cell in row:
            print(cell.value)

def open_excel(file_name):
    wb = op.load_workbook(file_path + file_name)
    ws = wb['test_sheet']

    for row in ws.rows:
        for cell in row :
            print('>>> {}'.format(cell.value))

    sheet = wb.active
    sheet['B1'] = 'success read'
    wb.save(file_path + file_name)

def update_cell_in_excel(file_name):
    wb = op.load_workbook(file_path + file_name)
    sheet = wb.active

    sheet['A1'] = 'success update data'
    wb.save(file_path + file_name)

def add_data_in_excel(file_name):
    wb = op.load_workbook(file_path + file_name)
    sheet = wb.active
    sheet.append(['add', 'new', 'data', 'by one commands'])
    wb.save(file_path + file_name)

def delete_data_in_excel(file_name, row):
    wb = op.load_workbook(file_path + file_name)
    sheet = wb.active
    sheet.delete_rows(row)
    wb.save(file_path + file_name)

#open_excel('test_sample.xlsx')
#update_cell_in_excel('test_sample.xlsx')
#read_excel('test_sample.xlsx')
